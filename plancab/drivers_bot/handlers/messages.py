import asyncio
import logging
import os
import qrcode
import tempfile
import re
import typing
import openpyxl
import os

from telethon import TelegramClient, events, Button, types, functions
from telethon.tl import patched

from ... import database, locales, shared, bot_utils, common

logger = logging.getLogger("drivers.messages")

add_new_temp_driver: typing.Dict[str, common.QuestionStep] = {
    "full_name": common.QuestionStep("kyc_full_name", common.ResponseType.TEXT),
    "vehicle_number": common.QuestionStep(
        "kyc_vehicle_number", common.ResponseType.TEXT
    ),
    "phone": common.QuestionStep("kyc_phone_number", common.ResponseType.TEXT),
    "vehicle_name": common.QuestionStep("kyc_vehicle_name", common.ResponseType.TEXT),
}

kyc_steps: typing.Dict[str, common.QuestionStep] = {
    "full_name": common.QuestionStep("kyc_full_name", common.ResponseType.TEXT),
    "vehicle_number": common.QuestionStep(
        "kyc_vehicle_number", common.ResponseType.TEXT
    ),
    "phone": common.QuestionStep("kyc_phone_number", common.ResponseType.TEXT),
    "vehicle_name": common.QuestionStep("kyc_vehicle_name", common.ResponseType.TEXT),
    "aadhar_card_photo": common.QuestionStep(
        "kyc_aadhar_card_photo", common.ResponseType.PHOTO
    ),
    "car_photo": common.QuestionStep("kyc_car_photo", common.ResponseType.PHOTO),
}


class UserPayload:
    kyc_step = "kyc_step"
    subscription_payment = "subscription_payment"
    edit_category_name = "edit_category_name"
    edit_category_cost_multiplier = "edit_category_cost_multiplier"
    add_category_name = "add_category_name"
    add_category_cost_multiplier = "add_category_cost_multiplier"
    feedback_comment = "feedback_comment"
    update_ride_driver = "update_ride_driver"


user_payload_binds = {}


def msg_payload_handler(data: str):
    def decorator(
            func: typing.Callable[
                [events.NewMessage.Event, database.User, bool, database.UserState],
                typing.Awaitable[None],
            ]
    ):
        global user_payload_binds
        user_payload_binds[data] = {"f": func}
        return func

    return decorator


commands_binds = {}


def command_handler(
        command: str,
        set_command=False,
        description_en: str = None,
        description_hi: str = None,
        description_kn: str = None,
        owner_only=False,
):
    def decorator(
            func: typing.Callable[
                [events.NewMessage.Event, database.User, bool, database.UserState],
                typing.Awaitable[None],
            ]
    ):
        global commands_binds
        commands_binds[command] = {
            "f": func,
            "set": set_command,
            "d_en": description_en,
            "d_hi": description_hi,
            "d_kn": description_kn,
            "owner_only": owner_only,
        }
        return func

    return decorator


@command_handler(
    "start",
    set_command=True,
    description_en="Sign up for driver",
    description_hi="ड्राइवर के लिए साइन अप करें",
    description_kn="ಡ್ರೈವರ್ ಗೆ ಸೈನ್ ಅಪ್ ಮಾಡಿ",
)
async def start_handler(
        event: events.NewMessage.Event,
        user: database.User,
        is_new_user: bool,
        state: database.UserState,
):
    state = await user.get_state(await event.client.get_peer_id("me"))
    state.payload = ""
    state.value = {}
    await state.save()

    if is_new_user:
        asyncio.create_task(
            bot_utils.send_user_to_privacy_share_group(event.client, event.get_sender())
        )
    await event.respond(file="driver_welcome.gif")
    # if ref := re.findall(r"start r(\d+)", event.raw_text):
    #     ref_by = int(ref[0])
    #     if ref_by != user.user_id:
    #         referer = await database.User.get_or_none(user_id=ref_by)
    #         referred = user
    #         if referer:
    #             exists = await database.ReferralPoint.get_or_none(
    #                 referer=referer, referred=referred
    #             )
    #             if not exists:
    #                 await database.ReferralPoint.create(
    #                     referer=referer,
    #                     referred=referred,
    #                 )
    driver = await user.get_driver()
    if driver.pending:
        await event.respond(user.loc.drivers_your_application_is_pending_text)
        return
    if is_new_user:
        # show a language selection menu
        await event.respond(
            user.loc.please_select_language_text,
            buttons=user.loc.select_language_buttons,
        )
        return
    if driver.confirmed:
        if driver.is_subscribed:
            await event.respond(
                user.loc.drivers_you_are_already_a_driver_has_sub_text.format(
                    sub_end_date=driver.subscription_until_tz_aw.strftime("%d %B %Y"),
                    invite_link=shared.config.driver_group_invite_link,
                ),
                buttons=user.loc.drivers_begin_payment_prolong_button,
            )
            return
        await event.respond(
            user.loc.drivers_you_are_already_a_driver_no_sub_text,
            buttons=user.loc.drivers_begin_payment_button,
        )
        return
    await event.respond(
        user.loc.drivers_please_complete_the_registration_text,
        buttons=user.loc.drivers_start_registration_button,
    )


# todo :  new code added
@command_handler("vendor")
async def change_to_vendor(
        event: events.NewMessage.Event,
        user: database.User,
        is_new_user: bool,
        state: database.UserState,
):
    driver = await user.get_driver()
    if driver.is_vendor:
        await event.respond(user.loc.you_are_already_vendor)
    elif driver.is_subscribed and driver.confirmed:
        buttons = [
            [Button.inline(text="Yes ✅", data="vendor_confirmation"), Button.inline(text="No ❌", data="start|clear")]]
        await event.respond(
            user.loc.confirm_vendor_upgrade, buttons=buttons
        )


@command_handler(
    "language",
    set_command=True,
    description_en="Select language",
    description_hi="भाषा चुनें",
    description_kn="ಭಾಷೆಯನ್ನು ಆಯ್ಕೆಮಾಡಿ",
)
async def language_handler(
        event: events.NewMessage.Event,
        user: database.User,
        is_new_user: bool,
        state: database.UserState,
):
    # show a language selection menu
    await event.respond(
        user.loc.please_select_language_text,
        buttons=user.loc.select_language_buttons,
    )


@command_handler(
    "referral",
    set_command=True,
    description_en="Refer a friend",
    description_hi="दोस्त को रेफर करें",
    description_kn="ಸ್ನೇಹಿತರಿಗೆ ಸೂಚಿಸಿ",
)
async def referral_handler(
        event: events.NewMessage.Event,
        user: database.User,
        is_new_user: bool,
        state: database.UserState,
):
    link = f"https://t.me/{shared.riders_bot_username}?start=r{user.user_id}"
    with tempfile.NamedTemporaryFile(suffix=".png") as tmp:
        qr = qrcode.make(link)
        qr.save(tmp, "PNG")
        tmp.seek(0)
        uploaded = await event.client.upload_file(tmp, file_name="qr.png")

    await event.respond(
        user.loc.drivers_referrals_info_text.format(
            referral_link=link,
            referral_count=await database.ReferralPoint.filter(referer=user).count(),
            redeemable_amount=await database.ReferralPoint.filter(
                referer=user, ride_id__isnull=False, point_used=False
            ).count(),
        ),
        file=uploaded,
    )


@command_handler(
    "link",
    set_command=True,
    description_en="Get driver group link",
    description_hi="ड्राइवर ग्रुप लिंक प्राप्त करें",
    description_kn="ಡ್ರೈವರ್ ಗ್ರೂಪ್ ಲಿಂಕ್ ಪಡೆಯಿರಿ",
)
async def link_handler(
        event: events.NewMessage.Event,
        user: database.User,
        is_new_user: bool,
        state: database.UserState,
):
    driver = await user.get_driver()
    if driver.confirmed:
        if driver.is_subscribed:
            await event.respond(
                user.loc.drivers_you_are_already_a_driver_has_sub_text.format(
                    sub_end_date=driver.subscription_until_tz_aw.strftime("%d %B %Y"),
                    invite_link=shared.config.driver_group_invite_link,
                ),
                buttons=user.loc.drivers_begin_payment_prolong_button,
            )
            return
        await event.respond(
            user.loc.drivers_you_are_already_a_driver_no_sub_text,
            buttons=user.loc.drivers_begin_payment_button,
        )
        return
    await event.respond(
        user.loc.drivers_please_complete_the_registration_text,
        buttons=user.loc.drivers_start_registration_button,
    )


# todo : new code added
@msg_payload_handler(UserPayload.update_ride_driver)
async def update_ride_driver_handler(
        event: events.NewMessage.Event,
        user: database.User,
        is_new_user: bool,
        state: database.UserState,
):
    question_key = state.value["step"]
    question_data = add_new_temp_driver[question_key]
    if not question_data.validate(event.message):
        await event.respond(user.loc.drivers_answer_incorrect_format_text)
        return
    if question_data.response_type == common.ResponseType.PHOTO:
        if event.message.file.size > 1024 * 1024 * 20:  # 20 MB
            await event.respond(user.loc.drivers_image_too_big_text)
            return
        state.payload = ""
        await state.save()
        m = await event.respond(user.loc.drivers_please_wait_downloading_text)

        if exists := state.value["data"].get(question_key):
            try:
                os.remove(exists)
            except OSError:
                pass

        try:
            file = await event.message.download_media("uploads")
        finally:
            state.payload = UserPayload.kyc_step
            await state.save()

        if not file:
            raise Exception("Failed to download media")

        file_ext = file.split(".")[-1]
        new_file = f"uploads/{question_key}_{user.user_id}.{file_ext}"
        if os.path.exists(new_file):
            os.remove(new_file)
        os.rename(file, new_file)

        asyncio.create_task(m.delete())

        state.value["data"][question_key] = new_file
    else:
        state.value["data"][question_key] = event.message.text

    add_new_temp_driver_keys = list(add_new_temp_driver.keys())
    is_last = question_key == add_new_temp_driver_keys[-1]

    if is_last:
        kyc_data = state.value["data"]
        driver = await database.Driver.create(user=user)
        driver.pending = False
        driver.vehicle_name = kyc_data["vehicle_name"]
        driver.vehicle_number = kyc_data["vehicle_number"]
        driver.phone = kyc_data["phone"]
        driver.full_name = kyc_data["full_name"]
        driver.vendor_id = state.value["vendor_id"]
        await driver.save()
        ride = await database.Ride.get(id=state.value["ride_id"])
        ride.driver = driver
        #ride.status = database.RideStatus.ACCEPTED_BY_DRIVER
        await ride.save()
        state.value = {}
        state.payload = ""
        await state.save()
        await shared.riders_bot.send_message(
            ride.user_id,
            user.loc.riders_user_drive_details_text.format(
                driver_details=bot_utils.get_max_user_info(
                    await shared.drivers_bot.get_entity(user.user_id)
                ),
                id=ride.id,
                driver_phone_number=driver.phone,
                driver_full_name=driver.full_name,
                vehicle_name=driver.vehicle_name,
                vehicle_plate_number=driver.vehicle_number,
            ),
        )

        return_ride = await database.Ride.filter(first_ride_id=ride.id)
        for i in return_ride:
            i.group_message_id = ride.group_message_id
            i.driver_id = ride.driver_id
            i.status = database.RideStatus.ACCEPTED_BY_DRIVER
            await i.save()

        await event.respond("Driver successfully updated ✔")
        return

    next_step_key = add_new_temp_driver_keys[add_new_temp_driver_keys.index(question_key) + 1]
    state.value["step"] = next_step_key
    next_step_data = add_new_temp_driver[next_step_key]

    buttons = [
        [Button.inline(user.loc.cancel_btntext, data="start|clear")],
    ]

    await state.save()
    await event.respond(
        next_step_data.get_text(user),
        buttons=buttons,
    )


@msg_payload_handler(UserPayload.kyc_step)
async def kyc_step_handler(
        event: events.NewMessage.Event,
        user: database.User,
        is_new_user: bool,
        state: database.UserState,
):
    question_key = state.value["step"]
    question_data = kyc_steps[question_key]
    if not question_data.validate(event.message):
        await event.respond(user.loc.drivers_answer_incorrect_format_text)
        return
    if question_data.response_type == common.ResponseType.PHOTO:
        if event.message.file.size > 1024 * 1024 * 20:  # 20 MB
            await event.respond(user.loc.drivers_image_too_big_text)
            return
        state.payload = ""
        await state.save()
        m = await event.respond(user.loc.drivers_please_wait_downloading_text)

        if exists := state.value["data"].get(question_key):
            try:
                os.remove(exists)
            except OSError:
                pass

        try:
            file = await event.message.download_media("uploads")
        finally:
            state.payload = UserPayload.kyc_step
            await state.save()

        if not file:
            raise Exception("Failed to download media")

        file_ext = file.split(".")[-1]
        new_file = f"uploads/{question_key}_{user.user_id}.{file_ext}"
        if os.path.exists(new_file):
            os.remove(new_file)
        os.rename(file, new_file)

        asyncio.create_task(m.delete())

        state.value["data"][question_key] = new_file
    else:
        state.value["data"][question_key] = event.message.text

    kyc_steps_keys = list(kyc_steps.keys())
    is_last = question_key == kyc_steps_keys[-1]

    prev_question_button = Button.inline(
        user.loc.prev_step_btntext,
        data=f"kyc_goto_step|" + kyc_steps_keys[kyc_steps_keys.index(question_key)],
    )

    if is_last:
        state.value["data"]["categories"] = []
        state.payload = ""
        await state.save()
        buttons = []
        all_enabled_categories = await database.CabCategory.filter(enabled=True)
        for category in all_enabled_categories:
            buttons.append(
                [
                    Button.inline(
                        "❌ " + category.name,
                        data=f"kyc_category|{category.id}",
                    )
                ]
            )
        buttons.append([prev_question_button])
        buttons.append([Button.inline(user.loc.cancel_btntext, data="start|clear")])
        await event.respond(
            user.loc.drivers_please_select_categories_text,
            buttons=buttons,
        )
        return

    next_step_key = kyc_steps_keys[kyc_steps_keys.index(question_key) + 1]
    state.value["step"] = next_step_key
    next_step_data = kyc_steps[next_step_key]

    buttons = [
        [prev_question_button],
        [Button.inline(user.loc.cancel_btntext, data="start|clear")],
    ]

    await state.save()
    await event.respond(
        next_step_data.get_text(user),
        buttons=buttons,
    )


def adjust_column_size_of_excel_sheet(sheet):
    max_width = 60
    dims = {}
    for row in sheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max(
                    (dims.get(cell.column_letter, 0), len(str(cell.value)))
                )
    for col, value in dims.items():
        sheet.column_dimensions[col].width = min(value + 3, max_width)


@command_handler(
    "export",
    set_command=True,
    description_en="Export database",
    owner_only=True,
)
async def export_handler(
        event: events.NewMessage.Event,
        user: database.User,
        is_new_user: bool,
        state: database.UserState,
):
    m = await event.respond(
        "Exporting database...",
    )
    categories = {x.id: x.name for x in await database.CabCategory.filter()}
    with tempfile.TemporaryDirectory() as tempdir:
        fname = os.path.join(tempdir, "export.xlsx")
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)
        users_sheet = workbook.create_sheet("Users")
        users_sheet.append(["id", "first_name", "last_name", "username", "language"])
        for user in await database.User.all():
            users_sheet.append(
                [
                    user.user_id,
                    user.first_name,
                    user.last_name,
                    user.username,
                    user.lang_code,
                ]
            )

        drivers_sheet = workbook.create_sheet("Drivers")
        drivers_sheet.append(
            [
                "driver id",
                "user id",
                "subscribed until",
                "full name",
                "vehicle number",
                "phone number",
                "vehicle name",
                "reg plate number",
                "confirmed",
            ]
        )
        for driver in await database.Driver.all():
            drivers_sheet.append(
                [
                    driver.id,
                    driver.user_id,
                    driver.subscription_until_tz_aw.strftime("%Y-%m-%d %H:%M:%S")
                    if driver.subscription_until_tz_aw
                    else None,
                    driver.full_name,
                    driver.vehicle_number,
                    driver.phone,
                    driver.vehicle_name,
                    driver.confirmed,
                ]
            )

        rides_sheet = workbook.create_sheet("rides")
        rides_sheet.append(
            [
                "id",
                "status",
                "driver id",
                "rider user id",
                "from",
                "to",
                "scheduled pickup time",
                "created at",
                "updated at",
                "price",
                "distance",
                "duration",
                "rider phone",
                "rider full name",
                "category",
                "rider's review rating",
                "rider's review text",
                "driver's review rating",
                "driver's review text",
            ]
        )
        for ride in await database.Ride.all():
            reviews_for_ride = await database.Review.filter(ride_id=ride.id)
            review_by_driver = None
            review_by_rider = None
            for review in reviews_for_ride:
                if review.by_driver:
                    review_by_driver = review
                else:
                    review_by_rider = review

            rides_sheet.append(
                [
                    ride.id,
                    ride.status,
                    ride.driver_id,
                    ride.user_id,
                    ride.from_text_address,
                    ride.to_text_address,
                    ride.pickup_time_scheduled_timezone_aware.strftime(
                        "%Y-%m-%d %H:%M:%S"
                    )
                    if ride.pickup_time_scheduled_timezone_aware
                    else None,
                    ride.created_at_timezone_aware.strftime("%Y-%m-%d %H:%M:%S")
                    if ride.created_at_timezone_aware
                    else None,
                    ride.updated_at_timezone_aware.strftime("%Y-%m-%d %H:%M:%S")
                    if ride.updated_at_timezone_aware
                    else None,
                    (ride.cost / 100) if ride.cost else None,
                    ride.distance,
                    ride.duration,
                    ride.phone_number,
                    ride.full_name,
                    categories[ride.category_id],
                    "⭐️" * review_by_rider.stars if review_by_rider else None,
                    review_by_rider.text if review_by_rider else None,
                    "⭐️" * review_by_driver.stars if review_by_driver else None,
                    review_by_driver.text if review_by_driver else None,
                ]
            )
        adjust_column_size_of_excel_sheet(users_sheet)
        adjust_column_size_of_excel_sheet(drivers_sheet)
        adjust_column_size_of_excel_sheet(rides_sheet)
        workbook.save(fname)
        asyncio.create_task(m.edit("Uploading..."))
        await event.respond(
            file=fname,
            force_document=True,
        )
        asyncio.create_task(m.delete())


@command_handler(
    "status",
    set_command=True,
    description_en="Get status about entity ",
    owner_only=True,
)
async def status_handler(
        event: events.NewMessage.Event,
        user: database.User,
        is_new_user: bool,
        state: database.UserState,
):
    message: patched.Message = event.message
    found = re.search(r"([udr])(\d+)", message.text)
    if not found:
        await event.respond(
            "Invalid syntax. Usage: /status <u+user_id | d+driver_id | r+ride_id>\n"
            "Example: `/status u1234567890` - shows info about user with id 1234567890\n"
            "`/status d1234567890` - shows info about driver with id 1234567890\n"
            "`/status r1234567890` - shows info about ride with id 1234567890\n"
        )
        return
    entity_type = found.group(1)
    user_id = int(found.group(2))

    match entity_type:
        case "u":
            target_user = await database.User.get_or_none(user_id=user_id)
            if not target_user:
                await event.respond("User not found")
                return
            rides_by_user = await database.Ride.filter(user_id=user_id)
            rides_by_state = {}
            for ride in rides_by_user:
                if ride.status not in rides_by_state:
                    rides_by_state[ride.status] = 0
                rides_by_state[ride.status] += 1
            rides_by_state = "\n".join(
                [f"`{k}`: {v}" for k, v in rides_by_state.items()]
            )
            if rides_by_state:
                rides_by_state = "\nRides by state:\n" + rides_by_state
            else:
                rides_by_state = ""
            referrer = await database.ReferralPoint.get_or_none(referred=target_user)
            referrer_text = ""
            if referrer:
                referrer_text = f"\nReferred by: {referrer.referer_id}\n"
            referral_count = await database.ReferralPoint.filter(
                referer=target_user
            ).count()
            redeemable_amount = await database.ReferralPoint.filter(
                referer=target_user, ride_id__isnull=False, point_used=False
            ).count()
            driver = await database.Driver.get_or_none(user_id=user_id)
            driver_text = "Driver id: no\n"
            if driver:
                driver_text = f"Driver id: `{driver.id}`"

            await event.respond(
                f"[User](tg://user?id={user_id}) {user_id}:\n"
                f"First name: {target_user.first_name}\n"
                f"Last name: {target_user.last_name or '❌'}\n"
                f"Username: {('@' + target_user.username) if target_user.username else '❌'}\n"
                f"Language code: {target_user.lang_code}\n"
                + rides_by_state
                + referrer_text
                + f"\nReferral count: {referral_count}\nRedeemable referral amount: {redeemable_amount}\n"
                  f"Banned: {'yes' if target_user.blocked else 'no'}\n" + driver_text,
                buttons=Button.inline(
                    "Ban" if not target_user.blocked else "Unban", f"ban|{user_id}"
                ),
            )
        case "d":
            target_driver = await database.Driver.get_or_none(id=user_id)
            if not target_driver:
                await event.respond(
                    "Driver not found. Make sure you are sending driver id, not user id"
                )
                return
            rides_by_driver = await database.Ride.filter(driver_id=user_id)
            total_earned = 0
            for ride in rides_by_driver:
                total_earned += ride.cost
            total_earned = round(total_earned / 100, 2)
            await event.respond(
                f"Driver id: `{target_driver.id}`\n"
                f"[User](tg://user?id={target_driver.user_id}) id: {target_driver.user_id}\n"
                f"Full name: {target_driver.full_name}\n"
                f"Vehicle number: {target_driver.vehicle_number}\n"
                f"Phone number: {target_driver.phone}\n"
                f"Vehicle name: {target_driver.vehicle_name}\n"
                f"KYC passed: {'yes' if target_driver.confirmed else 'no'}\n"
                f"Subscribed: {'no' if not target_driver.is_subscribed else ('until' + target_driver.subscription_until_tz_aw.strftime('%Y-%m-%d %H:%M:%S'))}\n"
                f"Total earned: {total_earned}\n"
                f"Total rides: {len(rides_by_driver)}\n"
            )
        case "r":
            target_ride = await database.Ride.get_or_none(id=user_id)
            if not target_ride:
                await event.respond("Ride not found")
                return
            await event.respond(await target_ride.full_admin_info())


@command_handler(
    "categories",
    description_en="Manage categories",
    description_hi="श्रेणियाँ प्रबंधित करें",
    description_kn="ವರ್ಗಗಳನ್ನು ನಿರ್ವಹಿಸಿ",
    owner_only=True,
)
async def categories_handler(
        event: events.NewMessage.Event,
        user: database.User,
        is_new_user: bool,
        state: database.UserState,
):
    all_categories = await database.CabCategory.all()
    buttons = []
    for category in all_categories:
        buttons.append(
            [
                Button.inline(
                    category.name,
                    data=f"lookup_category|{category.id}",
                )
            ]
        )
    buttons.append([Button.inline("Add new category", data="add_category")])
    await event.respond(
        "Click on a category to manage it",
        buttons=buttons,
    )


@msg_payload_handler(UserPayload.subscription_payment)
async def subscription_payment_handler(
        event: events.NewMessage.Event,
        user: database.User,
        is_new_user: bool,
        state: database.UserState,
):
    message: patched.Message = event.message
    if not (
            message.photo
            or (message.document and message.document.mime_type.startswith("image/"))
    ):
        await event.respond(user.loc.drivers_please_send_a_valid_image_text)
        return
    discount = state.value.get("discount")
    after_discount = state.value.get("after_discount")
    await shared.drivers_bot.send_message(
        shared.config.owner_id,
        f"New subscription payment from {bot_utils.get_max_user_info(await event.get_sender())}\n\n"
        f"Discount: {discount}%\n"
        f"Required amount: {after_discount}",
        file=message.media,
        buttons=[
            Button.inline(
                "✅ Accept",
                data=f"accept_subscription_payment|{user.user_id}|{message.id}|{discount}",
            ),
            Button.inline(
                "❌ Reject",
                data=f"reject_subscription_payment|{user.user_id}|{message.id}",
            ),
        ],
    )
    state.payload = ""
    state.value = {}
    await state.save()
    await event.respond(user.loc.drivers_payment_proof_sent_text)


@msg_payload_handler(UserPayload.edit_category_name)
async def edit_category_name_handler(
        event: events.NewMessage.Event,
        user: database.User,
        is_new_user: bool,
        state: database.UserState,
):
    if not event.message.text:
        raise Exception("No text")
    category = await database.CabCategory.get(id=state.value["category_id"])
    category.name = event.message.text
    await category.save()
    state.payload = ""
    state.value = {}
    await state.save()
    await event.respond(
        f"Category id: {category.id}\n"
        f"Name: {category.name}\n"
        f"Cost multiplier: {category.cost_multiplier}\n"
        f"Enabled: {category.enabled}\n",
        buttons=[
            [Button.inline("Edit name", data=f"edit_category_name|{category.id}")],
            [
                Button.inline(
                    "Edit cost multiplier",
                    data=f"edit_category_cost_multiplier|{category.id}",
                )
            ],
            [
                Button.inline(
                    "Toggle enabled", data=f"lookup_category|{category.id}|toggle"
                )
            ],
            [Button.inline("Back", data="back_to_categories")],
        ],
    )


@msg_payload_handler(UserPayload.edit_category_cost_multiplier)
async def edit_category_cost_multiplier_handler(
        event: events.NewMessage.Event,
        user: database.User,
        is_new_user: bool,
        state: database.UserState,
):
    if not event.message.text:
        raise Exception("No text")
    category = await database.CabCategory.get(id=state.value["category_id"])
    try:
        category.cost_multiplier = float(event.message.text)
    except ValueError:
        await event.respond("Invalid number")
        return
    await category.save()
    state.payload = ""
    state.value = {}
    await state.save()
    await event.respond(
        f"Category id: {category.id}\n"
        f"Name: {category.name}\n"
        f"Cost multiplier: {category.cost_multiplier}\n"
        f"Enabled: {category.enabled}\n",
        buttons=[
            [Button.inline("Edit name", data=f"edit_category_name|{category.id}")],
            [
                Button.inline(
                    "Edit cost multiplier",
                    data=f"edit_category_cost_multiplier|{category.id}",
                )
            ],
            [
                Button.inline(
                    "Toggle enabled", data=f"lookup_category|{category.id}|toggle"
                )
            ],
            [Button.inline("Back", data="back_to_categories")],
        ],
    )


@msg_payload_handler(UserPayload.add_category_name)
async def add_category_name_handler(
        event: events.NewMessage.Event,
        user: database.User,
        is_new_user: bool,
        state: database.UserState,
):
    if not event.message.text:
        raise Exception("No text")
    state.payload = UserPayload.add_category_cost_multiplier
    state.value = {"category_name": event.message.text}
    await state.save()
    await event.respond("Send me the cost multiplier")


@msg_payload_handler(UserPayload.add_category_cost_multiplier)
async def add_category_cost_multiplier_handler(
        event: events.NewMessage.Event,
        user: database.User,
        is_new_user: bool,
        state: database.UserState,
):
    if not event.message.text:
        raise Exception("No text")
    try:
        cost_multiplier = float(event.message.text)
    except ValueError:
        await event.respond("Invalid number")
        return
    category = await database.CabCategory.create(
        name=state.value["category_name"],
        cost_multiplier=cost_multiplier,
        website_code_name=state.value["category_name"].lower().replace(" ", "_"),
    )
    state.payload = ""
    state.value = {}
    await state.save()
    await event.respond(
        f"Category id: {category.id}\n"
        f"Name: {category.name}\n"
        f"Cost multiplier: {category.cost_multiplier}\n"
        f"Enabled: {category.enabled}\n",
        buttons=[
            [Button.inline("Edit name", data=f"edit_category_name|{category.id}")],
            [
                Button.inline(
                    "Edit cost multiplier",
                    data=f"edit_category_cost_multiplier|{category.id}",
                )
            ],
            [
                Button.inline(
                    "Toggle enabled", data=f"lookup_category|{category.id}|toggle"
                )
            ],
            [Button.inline("Back", data="back_to_categories")],
        ],
    )


@msg_payload_handler(UserPayload.feedback_comment)
async def feedback_comment_handler(
        event: events.NewMessage.Event,
        user: database.User,
        is_new_user: bool,
        state: database.UserState,
):
    message: patched.Message = event.message
    ride_id = state.value["feedback_ride_id"]
    ride = await database.Ride.get(id=ride_id)
    rating = state.value["feedback_rating"]
    await database.Review.create(
        stars=rating,
        text=message.text or "",
        from_user=user,
        to_user=await ride.user,
        ride=ride,
        by_driver=True,
    )
    asyncio.create_task(event.client.delete_messages(event.chat_id, state.value["feedback_delete_msg"]))
    state.payload = ""
    state.value = {}
    await state.save()
    await event.respond(user.loc.thanks_for_feedback_text)


async def message_handler(event: events.NewMessage.Event):
    global user_payload_binds, commands_binds
    user, is_new_user = await database.User.get_or_create(user_id=event.sender_id)
    asyncio.create_task(user.mb_update_data(event.message.get_sender()))
    if user.blocked:
        await event.respond(user.loc.you_are_blocked)
        raise events.StopPropagation
    if is_new_user:
        user_entity: types.User = await event.client.get_entity(event.sender_id)
        if user_entity.lang_code in locales.locales:
            logger.debug(
                f"Setting language for {user.user_id} to {user_entity.lang_code} - thanks telegram!"
            )
            user.lang_code = user_entity.lang_code
            await user.save()
    bot_id = await event.client.get_peer_id("me")
    state, _ = await database.UserState.get_or_create(user=user, bot_id=bot_id)
    text = event.raw_text
    logger.debug(
        f"Text query [u {user.user_id}] {{{state}}}: {text or type(event.message.media)}"
    )
    if text.startswith("/"):
        command = text.split(" ")[0][1:].split("@")[0]
        if command in commands_binds:
            try:
                if not commands_binds[command]['owner_only']:
                    await commands_binds[command]["f"](event, user, is_new_user, state)
                else:
                    if user.user_id == shared.config.owner_id:
                        await commands_binds[command]["f"](event, user, is_new_user, state)
            except Exception as e:
                logger.exception(e)
                await event.respond(user.loc.an_error_occurred)
        else:
            if not state.payload:
                logger.warning(f"Unknown command: {command}")
                await event.respond(user.loc.unknown_command)
    else:
        if state.payload:
            if state.payload in user_payload_binds:
                try:
                    await user_payload_binds[state.payload]["f"](
                        event, user, is_new_user, state
                    )
                except Exception as e:
                    logger.exception(e)
                    await event.respond(user.loc.an_error_occurred)
            else:
                logger.error(f"Unknown payload: {state.payload}")
                await event.respond(user.loc.unknown_payload)


async def set_commands(bot: TelegramClient):
    global commands_binds
    commands = {
        "en": [],
        "hi": [],
        "kn": [],
    }
    owner_only_commands = {
        "en": [],
        "hi": [],
        "kn": [],
    }

    for command, data in commands_binds.items():
        if data["d_en"]:
            owner_only_commands["en"].append(
                types.BotCommand(
                    command=command,
                    description=data["d_en"],
                )
            )
        if data["d_hi"]:
            owner_only_commands["hi"].append(
                types.BotCommand(
                    command=command,
                    description=data["d_hi"],
                )
            )
        if data["d_kn"]:
            owner_only_commands["kn"].append(
                types.BotCommand(
                    command=command,
                    description=data["d_kn"],
                )
            )
        if not data["owner_only"]:
            if data["d_en"]:
                commands["en"].append(
                    types.BotCommand(
                        command=command,
                        description=data["d_en"],
                    )
                )
            if data["d_hi"]:
                commands["hi"].append(
                    types.BotCommand(
                        command=command,
                        description=data["d_hi"],
                    )
                )
            if data["d_kn"]:
                commands["kn"].append(
                    types.BotCommand(
                        command=command,
                        description=data["d_kn"],
                    )
                )

    async def set_cmds(lc, com, scope, message):
        try:
            await bot(
                functions.bots.SetBotCommandsRequest(
                    commands=com,
                    scope=scope,
                    lang_code=lc,
                )
            )
        except Exception as e:
            logger.exception(e)
        logger.info(message)

    for lang_code, cmds in commands.items():
        if not cmds:
            continue
        asyncio.create_task(
            set_cmds(
                lang_code,
                cmds,
                types.BotCommandScopeUsers(),
                f"Set {len(cmds)} commands for {lang_code}",
            )
        )

    try:
        owner_peer = await bot.get_input_entity(shared.config.owner_id)
    except ValueError:
        logger.warning("Owner peer not found")
        return
    for lang_code, cmds in owner_only_commands.items():
        if not cmds:
            continue
        asyncio.create_task(
            set_cmds(
                lang_code,
                cmds,
                types.BotCommandScopePeer(peer=owner_peer),
                f"Set {len(cmds)} owner only commands for {lang_code}",
            )
        )


async def init_msg(bot: TelegramClient, parent_logger=None):
    global logger
    if parent_logger:
        logger = parent_logger.getChild("messages")
    logger.info("Setting up messages handler")
    bot.add_event_handler(message_handler, events.NewMessage(func=lambda e: e.is_private))
    logger.info("Setting up commands")
    await set_commands(bot)
    logger.info("Messages handler setup complete")
